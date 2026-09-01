from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/home/user/widgetdemo/a11y-audit/TOT_Dev_Fix_Sheet.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Dev Fix List"

FONT = "Arial"
NAVY = "164291"
INK = "171717"
GREY = "525252"

HDR_FILL = PatternFill("solid", fgColor=NAVY)
TOKEN_FILL = PatternFill("solid", fgColor="DCE8F7")
CODE_FILL = PatternFill("solid", fgColor="E8F5E9")
CHECK_FILL = PatternFill("solid", fgColor="FFF8E1")
GUARD_FILL = PatternFill("solid", fgColor="F5F5F5")
BAND = PatternFill("solid", fgColor="FAFAFA")

thin = Side(style="thin", color="D4D4D4")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["#", "Layer", "What to change", "Find / current", "Replace / new",
           "Where", "WCAG", "Contrast", "Effort", "Status", "Notes"]

# layer, what, find, replace, where, wcag, contrast, effort, notes
ROWS = [
    ("TOKEN", "Body / muted grey text",
     "#A3A3A3", "#767676",
     "Tailwind classes: text-[#A3A3A3], text-[#a3a3a3]",
     "1.4.3", "2.52:1 -> 4.54:1", "1 hr",
     "174 elements across 31 pages. Case-insensitive find and replace. Code-only: the Figma token is already correct at #525252, this hex was hardcoded in markup and never came from the token. Highest return item on this sheet."),

    ("TOKEN", "Functional border (inputs, control boundaries)",
     "#e5e5e5", "#737373",
     "--border-primary",
     "1.4.11", "1.26:1 -> 4.74:1", "30 min",
     "Functional borders only. Decorative dividers stay light and are exempt. Already changed in Figma (06-border/primary)."),

    ("TOKEN", "Strong border",
     "#a3a3a3", "#737373",
     "--border-strong",
     "1.4.11", "2.52:1 -> 4.74:1", "10 min",
     "Already correct in Figma (06-border/strong). Sync the code to match."),

    ("TOKEN", "Button hover fill",
     "#4398d4", "#2a6fa8",
     "--action-primary-hover",
     "1.4.3", "3.14:1 -> 5.33:1", "10 min",
     "Measured against the white label. Labels are 16px Semibold / 14px Bold / 12px Bold, none of which count as WCAG large text, so all need the full 4.5:1. Already correct in Figma."),

    ("TOKEN", "Focus ring",
     "browser default black", "#2a6fa8",
     "--focus-ring",
     "1.4.11", "2.22:1 -> 5.33:1", "included in code fix 12",
     "5.33:1 is measured against the page, which is why outline-offset is required. Without the offset the ring sits on the button and drops to 1.77:1."),

    ("TOKEN", "Focus ring, destructive buttons",
     "#f49898", "#991b1b",
     "--focus-ring-destructive",
     "1.4.11", "2.14:1 -> 8.31:1", "10 min",
     "Delete / cancel actions need their own ring. Already changed in Figma across all 3 destructive variants."),

    ("TOKEN", "Disabled text",
     "#d4d4d4", "#525252",
     "--text-disabled",
     "1.4.3 (exempt)", "1.48:1 -> 5.27:1", "10 min",
     "Disabled controls are technically exempt from 1.4.3. Do it anyway, they are currently unreadable. #525252 matches what the Figma Button and Input components already use."),

    ("TOKEN", "Badge, blue",
     "#3b82f6", "#2563eb",
     "badge background, white label",
     "1.4.3", "3.68:1 -> 5.17:1", "15 min",
     "Flat background, so this ratio is reliable."),

    ("TOKEN", "Badge, green",
     "#16a34a", "#15803d",
     "badge background, white label",
     "1.4.3", "3.30:1 -> 5.02:1", "15 min",
     "#16a34a stays fine for icons (3:1 threshold). Change it only where it is text."),

    ("TOKEN", "Badge, pink",
     "#ec4899", "#be185d",
     "badge background, white label",
     "1.4.3", "3.53:1 -> 6.04:1", "15 min",
     ""),

    ("TOKEN", "Badge, red label",
     "#fce6e7", "#ffffff",
     "label colour on the red badge",
     "1.4.3", "4.08:1 -> 4.87:1", "5 min",
     "The red background #e30910 itself is fine at 4.87:1, do not change it."),

    ("TOKEN", "Input field colours",
     "no component existed", "see Notes",
     "form field styles",
     "1.4.11, 1.4.3", "all pass", "30 min",
     "Build against Input- Master (Figma, COMPONENTS page). Border #737373 4.74:1 / focus ring #164291 9.45:1 drawn 2px OUTSIDE / error border #c2070d 6.31:1 / error text #9e1525 8.11:1 / placeholder #525252 7.81:1 / disabled #525252 on #fafafa 7.49:1. Field height 48px, which clears the 24x24 target minimum."),

    ("CODE", "Focus-visible ring, site-wide",
     "no custom focus style", ":focus-visible { outline: 2px solid #2a6fa8; outline-offset: 2px; }",
     "global stylesheet",
     "1.4.11, 2.4.7", "2.22:1 -> 5.33:1", "1 hr",
     "A ring DOES render today, it is the browser default black outline, which measures 2.22:1 against the #154291 buttons. Then grep for outline:none, outline:0, focus:outline-none and remove any with no replacement. Add .btn-destructive:focus-visible { outline-color: #991b1b; }"),

    ("CODE", "Skip to main content link",
     "first Tab stop is the cookie banner", '<a href="#main" class="skip-link">Skip to main content</a>',
     "every page, first element in the DOM",
     "2.4.1", "n/a", "1 hr",
     'Pair with <main id="main" tabindex="-1">. CSS: position:absolute; left:-9999px; z-index:9999; and .skip-link:focus { left:8px; top:8px; }. Must be before the cookie banner in the DOM, and check the banner does not trap or steal focus, or the link is useless.'),

    ("CODE", "Landmark elements",
     "generic divs", '<header role="banner"> <nav aria-label="Main"> <main id="main"> <footer role="contentinfo">',
     "site shell",
     "1.3.1", "n/a", "2 hrs",
     "Needed for the skip link to land anywhere useful."),

    ("CODE", "Horizontal scrollers reachable by keyboard",
     ".overflow-x-auto with no tabindex", 'tabindex="0" role="region" aria-label="..."',
     "20 regions across 14 pages",
     "2.1.1", "n/a", "2 hrs",
     "Give each region a label describing its own content, do not reuse one string."),

    ("CODE", "Route maps need a text alternative",
     "pannable map only", "stop and route data as real text or a table",
     "route map pages",
     "1.1.1, 2.1.1", "n/a", "varies",
     "A pannable map is unusable by a screen reader however focusable it is. If the stop data already exists as text elsewhere on the page, this is already done."),

    ("CODE", "Heading levels",
     "<h3> with no <h2> above it", "<h2>",
     "4 headings across 2 pages",
     "1.3.1", "n/a", "30 min",
     "Change the tag only, keep the Tailwind classes so nothing moves visually."),

    ("CODE", "Duplicate alt text",
     'alt="Kensington Palace" / alt="Tower of London"', 'alt=""',
     "attractions listing, 2 images on 1 page",
     "1.1.1", "n/a", "10 min",
     "The adjacent heading already names them, so they are announced twice today."),

    ("CODE", "Decorative illustrations",
     "alt text or missing attribute", 'alt=""',
     "12 illustrations",
     "1.1.1", "n/a", "15 min",
     'Confirmed decorative by the designer. alt="" and a MISSING alt attribute are not the same thing: an omitted attribute makes screen readers announce the filename. It must be present and empty.'),

    ("CODE", "Tap target hit areas",
     "icon buttons below minimum", ".icon-btn { min-width:44px; min-height:44px; display:grid; place-items:center; }",
     "~30 icon buttons and carousel arrows",
     "2.5.8", "n/a", "4 hrs",
     "Grow the hit area, not the icon, so there is no visual change. 24x24 is the AA floor, 44x44 is better on mobile."),

    ("CODE", "Form error markup",
     "unknown, inspect first", '<label for> + aria-invalid="true" + aria-describedby + role="alert"',
     "all forms",
     "3.3.1, 4.1.2", "n/a", "0-4 hrs",
     "The scan produced no evidence either way, so this may already be correct. Submit each form with invalid data and inspect before changing anything. Never signal an error by colour alone, keep the icon and the text."),

    ("GUARD", "Brand accent stays as it is",
     "#9cd1f3 / #63b5e8", "no change",
     "decorative washes and illustration fills",
     "1.4.11 (exempt)", "1.64:1 / 2.26:1", "0",
     "Confirmed decorative by the designer, so exempt. If either value is ever applied to an icon, link or active state, darken it to #3f8fc4 (3.10:1) first."),

    ("GUARD", "Alert borders stay as they are",
     "#86efac #fcd34d #f49898 #86b0e3", "no change",
     "success / warning / error / info alerts",
     "1.4.1, 1.4.11 (exempt)", "all below 3:1", "0",
     "Exempt ONLY because each alert also carries a distinct icon and a text label naming its type, so colour is reinforcement rather than the sole signal. Strip the icon or the label and these become real failures."),

    ("GUARD", "Do not change hero heading colours",
     "reported as failing", "no change without verifying",
     "hero headings, text over photos",
     "1.4.3", "unproven", "0",
     "~214 flagged rows are unproven: the contrast scanner could not resolve background images, so white text on hero photos was wrongly reported as failing. Run the verifier first (see check 26). For text on imagery, add a scrim rather than changing the text colour."),

    ("CHECK", "Settle the unproven contrast rows",
     "~214 rows flagged, unverified", "node scripts/07_verify_backdrops.js",
     "a11y-audit/local-scan",
     "1.4.3", "n/a", "30 min",
     "Sorts every ambiguous row into REAL_FAIL / PASSES / sits-on-an-image. Fix only REAL_FAIL. Skipping this means either fixing things that are not broken or missing ones that are. Do this before scoping the colour work."),

    ("CHECK", "Reflow at 200% and 400% zoom",
     "never validly tested", "manual test",
     "homepage, a PDP, a route map, a form",
     "1.4.10", "n/a", "2 hrs",
     "The scripted check reported 69/69 pages failing, which indicts the test rather than the site: it used CSS zoom, which is not how browser zoom works. At 400% (320px viewport) content must reflow to one column with no horizontal scrollbar. Commonly fails, so test it early."),

    ("CHECK", "Keyboard trap",
     "one page showed focus stuck", "manual test",
     "whole site, Tab from the address bar",
     "2.1.2", "n/a", "1 hr",
     "Level A, and the most severe failure class if real. Modals must close on Escape and return focus to their trigger."),

    ("CHECK", "Focus not obscured by the sticky header",
     "untested", "manual test",
     "every page with the sticky header",
     "2.4.11", "n/a", "1 hr",
     "New in WCAG 2.2 and a high-suspicion item here. Tab down the page and watch whether the focused element slides under the header."),

    ("CHECK", "Screen reader pass",
     "never performed", "NVDA (Windows) + VoiceOver (iOS)",
     "booking, search, contact journeys",
     "multiple", "n/a", "1 day",
     "No tool covers this. Scanning finds roughly 30-40% of WCAG issues, so a quiet scanner is not conformance."),

    ("CHECK", "Stop regressions in CI",
     "no automated gate", "npx playwright test a11y.spec.ts",
     "CI, on every PR",
     "n/a", "n/a", "2 hrs",
     "npm i -D @axe-core/playwright, then fail the build on any violation. Run it against a preview deploy. The full 69-page crawl takes minutes. Pipeline is in a11y-audit/local-scan/."),
]

# ---- header block ----
ws["A1"] = "Accessibility fix list — theoriginaltour.com"
ws["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
ws["A2"] = ("WCAG 2.2 Level AA. Every value below is final — nothing is waiting on a designer. "
            "Work TOKEN rows first (fastest, widest reach), then CODE, then the CHECK rows before shipping. "
            "GUARD rows are things NOT to change, and why.")
ws["A2"].font = Font(name=FONT, size=10, color=GREY)
ws["A3"] = ('Scope: all 69 pages, desktop 1440 and mobile 390, scanned 31 Aug 2026. Contrast ratios computed with the '
            'WCAG relative-luminance formula from source values and cross-checked against the Figma tokens.')
ws["A3"].font = Font(name=FONT, size=10, italic=True, color=GREY)

first_data_row = 8
last_data_row = first_data_row + len(ROWS) - 1

# Deliberately no formulas in this workbook. It is a checklist, not a model, and
# LibreOffice could not complete a recalculation in the build environment — so a
# progress counter would have shipped with no cached value and no way to verify it.
ws["A5"] = "Tracking:"
ws["A5"].font = Font(name=FONT, size=10, bold=True)
ws["B5"] = "Set column J as you go. Use the filter on row 7 to show only Not started, or only TOKEN rows."
ws["B5"].font = Font(name=FONT, size=10, color=NAVY)

HDR_ROW = 7
for c, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=HDR_ROW, column=c, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = BOX
ws.row_dimensions[HDR_ROW].height = 26

LAYER_FILL = {"TOKEN": TOKEN_FILL, "CODE": CODE_FILL, "GUARD": GUARD_FILL, "CHECK": CHECK_FILL}

for i, (layer, what, find, repl, where, wcag, contrast, effort, notes) in enumerate(ROWS):
    r = first_data_row + i
    values = [i + 1, layer, what, find, repl, where, wcag, contrast, effort, "Not started", notes]
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BOX
        if i % 2 == 1:
            cell.fill = BAND
    ws.cell(row=r, column=2).fill = LAYER_FILL[layer]
    ws.cell(row=r, column=2).font = Font(name=FONT, size=9, bold=True, color=INK)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="top")
    # monospace-ish emphasis on the actual values to change
    for col in (4, 5):
        ws.cell(row=r, column=col).font = Font(name="Consolas", size=9, color=INK)
    ws.cell(row=r, column=3).font = Font(name=FONT, size=10, bold=True, color=INK)
    ws.cell(row=r, column=11).font = Font(name=FONT, size=9, color=GREY)
    ws.row_dimensions[r].height = 58

dv = DataValidation(type="list", formula1='"Not started,In progress,Blocked,Done,N/A"',
                    allow_blank=False, showErrorMessage=True)
dv.errorTitle = "Pick a status"
dv.error = "Choose one of the listed values."
ws.add_data_validation(dv)
dv.add(f"J{first_data_row}:J{last_data_row}")

widths = {"A": 4, "B": 8, "C": 30, "D": 24, "E": 30, "F": 26, "G": 14, "H": 17, "I": 11, "J": 12, "K": 62}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = f"A{first_data_row}"
ws.auto_filter.ref = f"A{HDR_ROW}:K{last_data_row}"

# ---- legend ----
lr = last_data_row + 2
ws.cell(row=lr, column=1, value="Legend").font = Font(name=FONT, size=11, bold=True, color=NAVY)
legend = [
    ("TOKEN", "A colour value. Change it once where it is defined and it propagates.", TOKEN_FILL),
    ("CODE", "A markup or CSS change in the components themselves.", CODE_FILL),
    ("GUARD", "Do NOT change this. Listed so it is not 'fixed' by mistake.", GUARD_FILL),
    ("CHECK", "Must be verified by a person before any AA claim. Tooling cannot settle it.", CHECK_FILL),
]
for j, (k, d, fill) in enumerate(legend):
    r = lr + 1 + j
    a = ws.cell(row=r, column=1, value=k)
    a.fill = fill
    a.font = Font(name=FONT, size=9, bold=True)
    a.border = BOX
    b = ws.cell(row=r, column=3, value=d)
    b.font = Font(name=FONT, size=10, color=GREY)

fr = lr + 6
ws.cell(row=fr, column=1, value="This sheet on its own will not make the site AA conformant.").font = Font(
    name=FONT, size=11, bold=True, color="92400E")
ws.cell(row=fr + 1, column=1, value=(
    "Automated checks reliably detect roughly 30-40% of WCAG issues. 27 success criteria are still untested — "
    "keyboard operation, screen-reader output, reflow, focus-not-obscured and error announcement only exist once built. "
    "A single unresolved failure anywhere breaks conformance, so no AA claim should be made until the CHECK rows are closed. "
    "If some items cannot close before launch, publish a PARTIAL conformance statement naming exactly what does not yet "
    "conform and when it will — an inaccurate blanket AA claim carries more legal exposure than an honest one.")
).font = Font(name=FONT, size=10, color=GREY)
ws.cell(row=fr + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=fr + 1, start_column=1, end_row=fr + 3, end_column=11)

wb.save(OUT)
print("written:", OUT, "rows:", len(ROWS))
